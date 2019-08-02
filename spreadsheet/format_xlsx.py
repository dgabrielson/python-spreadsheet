#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Reading and writing XLSX Excel spreadsheets
"""
from __future__ import print_function, unicode_literals

from io import BytesIO

from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.writer.excel import save_workbook

import spreadsheet  # for spreadsheet.Formula


def reader(filename, fileobj, **kwargs):
    """
    ``fileobj`` should be in binary, and at the beginning of the stream.

    """
    # load_workbook backs onto zipfile.ZipFile, which supports file objects or filenames.
    book = load_workbook(fileobj or filename, **kwargs)
    first_sheet_name = book.get_sheet_names()[0]
    sheet = book.get_sheet_by_name(first_sheet_name)
    results = []
    M = sheet.get_highest_row()
    N = sheet.get_highest_column()
    for i in range( M ):
        row = []
        for j in range( N ):
            cell = sheet.cell(row=i, column=j)
            row.append( cell.value )
        while row and row[-1] is None:
            row.pop()
        results.append( row )   # this processes formulas.
    return results




def writer(data, **kwargs):
    """
    Formulas are easy with this one, just do '=F1+F2'
    """
    def handle_formula(f):
        return "{}".format(f)


    book = Workbook(**kwargs)
    sheet = book.worksheets[0]
    sheet.title = u'Sheet 1'

    for i in range(len(data)):
        row = data[i]
        for j in range(len(row)):
            item = row[j] if not isinstance(row[j], spreadsheet.Formula) else handle_formula(row[j])
            if item is not None:
                cell = sheet.cell(row=i, column=j)
                cell.value = "{}".format(item)

    result = BytesIO()
    save_workbook(book, result) # Known to cause problems with Python 2.7, at this time.
    return result.getvalue()


##
